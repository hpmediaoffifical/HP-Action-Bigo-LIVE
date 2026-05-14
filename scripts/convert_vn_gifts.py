"""Convert file Excel Việt Nam → JSON. Auto-detect cột ID/Tên/Link/Giá theo header."""
import openpyxl, json, sys, time, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = r'C:\Users\NCPC\Downloads\Qùa NPC_14052026.xlsx'
DST = r'C:\Users\NCPC\Desktop\BIGO Action\config\vietnam-gifts.json'

wb = openpyxl.load_workbook(SRC, data_only=True)

def find_col(headers, *keywords):
    for idx, h in enumerate(headers):
        if h is None:
            continue
        s = str(h).strip().upper()
        if any(k in s for k in keywords):
            return idx
    return None

all_gifts = []
seen_ids = set()
sheet_summary = {}

for sn in wb.sheetnames:
    ws = wb[sn]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
    # Tìm header row (chứa "ID QUÀ" hoặc "ID")
    hdr_idx = None
    for i, r in enumerate(rows[:5]):
        if r and any(c and 'ID' in str(c).upper() for c in r):
            hdr_idx = i
            break
    if hdr_idx is None:
        continue
    headers = rows[hdr_idx]
    c_id = find_col(headers, 'ID QUÀ', 'ID')
    c_name = find_col(headers, 'TÊN QUÀ', 'TÊN', 'NAME')
    c_url = find_col(headers, 'ẢNH', 'LINK', 'URL', 'IMAGE')
    c_dia = find_col(headers, 'ĐƠN GIÁ', 'KIM CƯƠNG', 'KC', 'DIAMOND', 'PRICE')

    count_sheet = 0
    for r in rows[hdr_idx + 1:]:
        if not r or all(c is None or str(c).strip() == '' for c in r):
            continue
        try:
            raw_id = r[c_id]
            if raw_id is None or str(raw_id).strip() == '':
                continue
            typeid = int(float(str(raw_id).strip()))
        except (ValueError, TypeError):
            continue
        if typeid in seen_ids:
            continue
        seen_ids.add(typeid)
        name = (str(r[c_name]).strip() if c_name is not None and r[c_name] is not None else '')
        url = (str(r[c_url]).strip() if c_url is not None and r[c_url] is not None else '')
        try:
            diamonds = int(float(str(r[c_dia]).strip())) if c_dia is not None and r[c_dia] is not None else None
        except (ValueError, TypeError):
            diamonds = None
        all_gifts.append({
            'typeid': typeid,
            'name': name,
            'img_url': url,
            'diamonds': diamonds,
            'sheet': sn,
        })
        count_sheet += 1
    sheet_summary[sn] = count_sheet

out = {
    'region': 'vn',
    'source': 'Qùa NPC_14052026.xlsx',
    'fetchedAt': int(time.time() * 1000),
    'sheetSummary': sheet_summary,
    'gifts': sorted(all_gifts, key=lambda g: g['typeid']),
}

os.makedirs(os.path.dirname(DST), exist_ok=True)
with open(DST, 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

print(f"Saved {DST}")
print(f"Total: {len(all_gifts)} unique gifts")
print(f"Sheets: {sheet_summary}")
print(f"Sample first 3:")
for g in out['gifts'][:3]:
    print(f"  typeid={g['typeid']}, name={g['name']}, diamonds={g['diamonds']}")
